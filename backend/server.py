from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, date
import os
import uuid
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = FastAPI()

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB connection
MONGO_URL = os.environ.get('MONGO_URL')
client = AsyncIOMotorClient(MONGO_URL)
db = client.painting_contractor_db

# Pydantic Models
class Site(BaseModel):
    site_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    owner_name: str
    owner_phone: str
    owner_email: Optional[str] = None
    location: str
    maps_link: Optional[str] = None
    start_date: str
    status: str = "Running"  # Running, Completed, On Hold
    created_at: str = Field(default_factory=lambda: datetime.now().isoformat())

    class Config:
        json_schema_extra = {
            "example": {
                "name": "Site A - House Painting",
                "owner_name": "John Doe",
                "owner_phone": "9876543210",
                "owner_email": "john@example.com",
                "location": "123 Main St, Mumbai",
                "maps_link": "https://maps.google.com",
                "start_date": "2025-08-01",
                "status": "Running"
            }
        }

class Material(BaseModel):
    material_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    unit: str  # bucket, piece, kg, liter
    rate_per_unit: float
    current_stock: float
    created_at: str = Field(default_factory=lambda: datetime.now().isoformat())

    class Config:
        json_schema_extra = {
            "example": {
                "name": "Asian Paint - White",
                "unit": "bucket",
                "rate_per_unit": 1200.0,
                "current_stock": 50.0
            }
        }

class Labour(BaseModel):
    labour_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    rate_per_day: float
    created_at: str = Field(default_factory=lambda: datetime.now().isoformat())

    class Config:
        json_schema_extra = {
            "example": {
                "name": "Rajesh Kumar",
                "rate_per_day": 800.0
            }
        }

class MaterialUsed(BaseModel):
    material_id: str
    material_name: str
    quantity: float
    rate_per_unit: float
    total_cost: float

class LabourUsed(BaseModel):
    labour_id: str
    labour_name: str
    count: int
    rate_per_day: float
    total_cost: float

class SiteDailyLog(BaseModel):
    log_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    site_id: str
    site_name: str
    log_date: str
    materials_used: List[MaterialUsed] = []
    labours_used: List[LabourUsed] = []
    notes: Optional[str] = None
    total_material_cost: float = 0.0
    total_labour_cost: float = 0.0
    total_cost: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now().isoformat())

class Overhead(BaseModel):
    overhead_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    site_id: str
    site_name: str
    date: str
    category: str  # Transport, Food, Scaffolding, Miscellaneous
    amount: float
    description: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now().isoformat())

# Helper function to serialize MongoDB documents
def serialize_doc(doc):
    if doc and '_id' in doc:
        doc.pop('_id')
    return doc

# SITES ROUTES
@app.get("/api/sites", response_model=List[Site])
async def get_sites():
    sites = await db.sites.find().to_list(length=None)
    return [serialize_doc(site) for site in sites]

@app.post("/api/sites", response_model=Site)
async def create_site(site: Site):
    site_dict = site.dict()
    await db.sites.insert_one(site_dict)
    return serialize_doc(site_dict)

@app.put("/api/sites/{site_id}", response_model=Site)
async def update_site(site_id: str, site: Site):
    site_dict = site.dict()
    site_dict['site_id'] = site_id
    result = await db.sites.replace_one({"site_id": site_id}, site_dict)
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Site not found")
    return serialize_doc(site_dict)

@app.delete("/api/sites/{site_id}")
async def delete_site(site_id: str):
    result = await db.sites.delete_one({"site_id": site_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Site not found")
    # Also delete related logs and overheads
    await db.site_daily_logs.delete_many({"site_id": site_id})
    await db.overheads.delete_many({"site_id": site_id})
    return {"message": "Site deleted successfully"}

# MATERIALS ROUTES
@app.get("/api/materials", response_model=List[Material])
async def get_materials():
    materials = await db.materials.find().to_list(length=None)
    return [serialize_doc(material) for material in materials]

@app.post("/api/materials", response_model=Material)
async def create_material(material: Material):
    material_dict = material.dict()
    await db.materials.insert_one(material_dict)
    return serialize_doc(material_dict)

@app.put("/api/materials/{material_id}", response_model=Material)
async def update_material(material_id: str, material: Material):
    material_dict = material.dict()
    material_dict['material_id'] = material_id
    result = await db.materials.replace_one({"material_id": material_id}, material_dict)
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    return serialize_doc(material_dict)

@app.delete("/api/materials/{material_id}")
async def delete_material(material_id: str):
    result = await db.materials.delete_one({"material_id": material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    return {"message": "Material deleted successfully"}

# LABOURS ROUTES
@app.get("/api/labours", response_model=List[Labour])
async def get_labours():
    labours = await db.labours.find().to_list(length=None)
    return [serialize_doc(labour) for labour in labours]

@app.post("/api/labours", response_model=Labour)
async def create_labour(labour: Labour):
    labour_dict = labour.dict()
    await db.labours.insert_one(labour_dict)
    return serialize_doc(labour_dict)

@app.put("/api/labours/{labour_id}", response_model=Labour)
async def update_labour(labour_id: str, labour: Labour):
    labour_dict = labour.dict()
    labour_dict['labour_id'] = labour_id
    result = await db.labours.replace_one({"labour_id": labour_id}, labour_dict)
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Labour not found")
    return serialize_doc(labour_dict)

@app.delete("/api/labours/{labour_id}")
async def delete_labour(labour_id: str):
    result = await db.labours.delete_one({"labour_id": labour_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Labour not found")
    return {"message": "Labour deleted successfully"}

# SITE DAILY LOGS ROUTES
@app.get("/api/site-logs", response_model=List[SiteDailyLog])
async def get_site_logs(site_id: Optional[str] = None):
    query = {"site_id": site_id} if site_id else {}
    logs = await db.site_daily_logs.find(query).sort("log_date", -1).to_list(length=None)
    return [serialize_doc(log) for log in logs]

@app.post("/api/site-logs", response_model=SiteDailyLog)
async def create_site_log(log: SiteDailyLog):
    log_dict = log.dict()
    
    # Calculate costs
    material_cost = sum(m['total_cost'] for m in log_dict['materials_used'])
    labour_cost = sum(l['total_cost'] for l in log_dict['labours_used'])
    log_dict['total_material_cost'] = material_cost
    log_dict['total_labour_cost'] = labour_cost
    log_dict['total_cost'] = material_cost + labour_cost
    
    # Update central inventory - reduce stock
    for material_used in log_dict['materials_used']:
        material = await db.materials.find_one({"material_id": material_used['material_id']})
        if material:
            new_stock = material['current_stock'] - material_used['quantity']
            await db.materials.update_one(
                {"material_id": material_used['material_id']},
                {"$set": {"current_stock": new_stock}}
            )
    
    await db.site_daily_logs.insert_one(log_dict)
    return serialize_doc(log_dict)

@app.put("/api/site-logs/{log_id}", response_model=SiteDailyLog)
async def update_site_log(log_id: str, log: SiteDailyLog):
    # First, get the old log to restore stock
    old_log = await db.site_daily_logs.find_one({"log_id": log_id})
    if not old_log:
        raise HTTPException(status_code=404, detail="Log not found")
    
    # Restore old stock
    for material_used in old_log['materials_used']:
        material = await db.materials.find_one({"material_id": material_used['material_id']})
        if material:
            new_stock = material['current_stock'] + material_used['quantity']
            await db.materials.update_one(
                {"material_id": material_used['material_id']},
                {"$set": {"current_stock": new_stock}}
            )
    
    # Update with new data
    log_dict = log.dict()
    log_dict['log_id'] = log_id
    
    # Calculate costs
    material_cost = sum(m['total_cost'] for m in log_dict['materials_used'])
    labour_cost = sum(l['total_cost'] for l in log_dict['labours_used'])
    log_dict['total_material_cost'] = material_cost
    log_dict['total_labour_cost'] = labour_cost
    log_dict['total_cost'] = material_cost + labour_cost
    
    # Reduce stock for new materials
    for material_used in log_dict['materials_used']:
        material = await db.materials.find_one({"material_id": material_used['material_id']})
        if material:
            new_stock = material['current_stock'] - material_used['quantity']
            await db.materials.update_one(
                {"material_id": material_used['material_id']},
                {"$set": {"current_stock": new_stock}}
            )
    
    await db.site_daily_logs.replace_one({"log_id": log_id}, log_dict)
    return serialize_doc(log_dict)

@app.delete("/api/site-logs/{log_id}")
async def delete_site_log(log_id: str):
    log = await db.site_daily_logs.find_one({"log_id": log_id})
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    
    # Restore stock
    for material_used in log['materials_used']:
        material = await db.materials.find_one({"material_id": material_used['material_id']})
        if material:
            new_stock = material['current_stock'] + material_used['quantity']
            await db.materials.update_one(
                {"material_id": material_used['material_id']},
                {"$set": {"current_stock": new_stock}}
            )
    
    await db.site_daily_logs.delete_one({"log_id": log_id})
    return {"message": "Log deleted successfully"}

# OVERHEADS ROUTES
@app.get("/api/overheads", response_model=List[Overhead])
async def get_overheads(site_id: Optional[str] = None):
    query = {"site_id": site_id} if site_id else {}
    overheads = await db.overheads.find(query).sort("date", -1).to_list(length=None)
    return [serialize_doc(overhead) for overhead in overheads]

@app.post("/api/overheads", response_model=Overhead)
async def create_overhead(overhead: Overhead):
    overhead_dict = overhead.dict()
    await db.overheads.insert_one(overhead_dict)
    return serialize_doc(overhead_dict)

@app.put("/api/overheads/{overhead_id}", response_model=Overhead)
async def update_overhead(overhead_id: str, overhead: Overhead):
    overhead_dict = overhead.dict()
    overhead_dict['overhead_id'] = overhead_id
    result = await db.overheads.replace_one({"overhead_id": overhead_id}, overhead_dict)
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Overhead not found")
    return serialize_doc(overhead_dict)

@app.delete("/api/overheads/{overhead_id}")
async def delete_overhead(overhead_id: str):
    result = await db.overheads.delete_one({"overhead_id": overhead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Overhead not found")
    return {"message": "Overhead deleted successfully"}

# REPORTS ROUTES
@app.get("/api/reports/site/{site_id}")
async def get_site_report(site_id: str):
    site = await db.sites.find_one({"site_id": site_id})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    
    logs = await db.site_daily_logs.find({"site_id": site_id}).to_list(length=None)
    overheads = await db.overheads.find({"site_id": site_id}).to_list(length=None)
    
    total_material_cost = sum(log['total_material_cost'] for log in logs)
    total_labour_cost = sum(log['total_labour_cost'] for log in logs)
    total_overhead_cost = sum(overhead['amount'] for overhead in overheads)
    grand_total = total_material_cost + total_labour_cost + total_overhead_cost
    
    return {
        "site": serialize_doc(site),
        "total_material_cost": total_material_cost,
        "total_labour_cost": total_labour_cost,
        "total_overhead_cost": total_overhead_cost,
        "grand_total": grand_total,
        "logs_count": len(logs),
        "overheads_count": len(overheads)
    }

@app.get("/api/reports/daily")
async def get_daily_report(date: Optional[str] = None):
    query = {"log_date": date} if date else {}
    logs = await db.site_daily_logs.find(query).to_list(length=None)
    
    total_cost = sum(log['total_cost'] for log in logs)
    
    return {
        "date": date or "All dates",
        "logs": [serialize_doc(log) for log in logs],
        "total_cost": total_cost
    }

@app.get("/api/reports/inventory")
async def get_inventory_report():
    materials = await db.materials.find().to_list(length=None)
    
    total_stock_value = sum(m['current_stock'] * m['rate_per_unit'] for m in materials)
    low_stock_items = [m for m in materials if m['current_stock'] < 5]
    
    return {
        "materials": [serialize_doc(m) for m in materials],
        "total_stock_value": total_stock_value,
        "low_stock_items": [serialize_doc(m) for m in low_stock_items]
    }

# EXCEL EXPORT ROUTES
@app.get("/api/export/site/{site_id}")
async def export_site_report(site_id: str):
    site = await db.sites.find_one({"site_id": site_id})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    
    logs = await db.site_daily_logs.find({"site_id": site_id}).sort("log_date", 1).to_list(length=None)
    overheads = await db.overheads.find({"site_id": site_id}).sort("date", 1).to_list(length=None)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Site Report"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Site Information
    ws['A1'] = "Site Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Site Name: {site['name']}"
    ws['A3'] = f"Owner: {site['owner_name']}"
    ws['A4'] = f"Location: {site['location']}"
    ws['A5'] = f"Status: {site['status']}"
    ws['A6'] = ""
    
    # Daily Logs Section
    row = 7
    ws[f'A{row}'] = "Daily Logs"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    headers = ['Date', 'Materials Cost', 'Labour Cost', 'Total Cost', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    total_material = 0
    total_labour = 0
    for log in logs:
        ws.cell(row=row, column=1, value=log['log_date'])
        ws.cell(row=row, column=2, value=log['total_material_cost'])
        ws.cell(row=row, column=3, value=log['total_labour_cost'])
        ws.cell(row=row, column=4, value=log['total_cost'])
        ws.cell(row=row, column=5, value=log.get('notes', ''))
        total_material += log['total_material_cost']
        total_labour += log['total_labour_cost']
        row += 1
    
    # Overheads Section
    row += 1
    ws[f'A{row}'] = "Overheads"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    headers = ['Date', 'Category', 'Amount', 'Description']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    total_overhead = 0
    for overhead in overheads:
        ws.cell(row=row, column=1, value=overhead['date'])
        ws.cell(row=row, column=2, value=overhead['category'])
        ws.cell(row=row, column=3, value=overhead['amount'])
        ws.cell(row=row, column=4, value=overhead.get('description', ''))
        total_overhead += overhead['amount']
        row += 1
    
    # Summary
    row += 1
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    ws[f'A{row}'] = "Total Material Cost:"
    ws[f'B{row}'] = total_material
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Total Labour Cost:"
    ws[f'B{row}'] = total_labour
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Total Overhead Cost:"
    ws[f'B{row}'] = total_overhead
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Grand Total:"
    ws[f'B{row}'] = total_material + total_labour + total_overhead
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'B{row}'].font = Font(size=12, bold=True, color="FF0000")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=site_report_{site['name']}.xlsx"}
    )

@app.get("/api/export/inventory")
async def export_inventory_report():
    materials = await db.materials.find().to_list(length=None)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws['A1'] = "Central Material Inventory Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = ""
    
    headers = ['Material Name', 'Unit', 'Rate per Unit', 'Current Stock', 'Stock Value']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 5
    total_value = 0
    for material in materials:
        stock_value = material['current_stock'] * material['rate_per_unit']
        ws.cell(row=row, column=1, value=material['name'])
        ws.cell(row=row, column=2, value=material['unit'])
        ws.cell(row=row, column=3, value=material['rate_per_unit'])
        ws.cell(row=row, column=4, value=material['current_stock'])
        ws.cell(row=row, column=5, value=stock_value)
        total_value += stock_value
        row += 1
    
    row += 1
    ws[f'D{row}'] = "Total Stock Value:"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'] = total_value
    ws[f'E{row}'].font = Font(bold=True, color="FF0000")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"}
    )

@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "message": "Painting Contractor API is running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)